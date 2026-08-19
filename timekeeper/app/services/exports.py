"""Report exporters — CSV, XLSX and PDF (FR-I-10).

US-06 AC-2: the exported totals must equal the totals shown on screen, so both
paths format through the same `render_value` helper.
"""

from __future__ import annotations

import csv
import io
from datetime import date, datetime

from . import timeutil as T


def render_value(value, column: dict, duration_format: str = "hm"):
    if value is None:
        return ""
    if column.get("type") == "duration":
        return T.format_duration(int(value), duration_format)
    if isinstance(value, bool):
        return "yes" if value else "no"
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def has_column_totals(report: dict) -> bool:
    """A totals row is only meaningful when at least one column has one; the
    live board, for instance, totals states rather than columns."""
    totals = report.get("totals") or {}
    return any(column["key"] in totals for column in report["columns"])


def totals_row(report: dict, duration_format: str) -> list:
    row = []
    for index, column in enumerate(report["columns"]):
        key = column["key"]
        if key in report.get("totals", {}):
            row.append(render_value(report["totals"][key], column, duration_format))
        elif index == 0:
            row.append("TOTAL")
        else:
            row.append("")
    return row


def to_csv(report: dict, duration_format: str = "hm", delimiter: str = ",") -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=delimiter, lineterminator="\r\n")
    writer.writerow([c["label"] for c in report["columns"]])
    for row in report["rows"]:
        writer.writerow(
            [render_value(row.get(c["key"]), c, duration_format) for c in report["columns"]]
        )
    if has_column_totals(report):
        writer.writerow(totals_row(report, duration_format))
    return buffer.getvalue().encode("utf-8-sig")


def to_xlsx(report: dict, duration_format: str = "hm") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = report["type"][:31]

    sheet.append([report.get("title", "Report")])
    sheet["A1"].font = Font(size=14, bold=True)
    meta = report.get("meta", {})
    sheet.append(
        [f"{key}: {value}" for key, value in meta.items() if not isinstance(value, dict)]
    )
    sheet.append([])

    header_row = sheet.max_row + 1
    sheet.append([c["label"] for c in report["columns"]])
    fill = PatternFill("solid", fgColor="1F3A5F")
    for cell in sheet[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row in report["rows"]:
        sheet.append(
            [render_value(row.get(c["key"]), c, duration_format) for c in report["columns"]]
        )
    if has_column_totals(report):
        sheet.append(totals_row(report, duration_format))
        for cell in sheet[sheet.max_row]:
            cell.font = Font(bold=True)

    for index, column in enumerate(report["columns"], start=1):
        width = max(len(str(column["label"])) + 2, 12)
        sheet.column_dimensions[get_column_letter(index)].width = min(width, 40)
    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def to_pdf(report: dict, duration_format: str = "hm", org_name: str = "") -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
        title=report.get("title", "Report"),
    )
    styles = getSampleStyleSheet()
    small = styles["BodyText"].clone("small")
    small.fontSize = 7
    small.leading = 8.5

    story = [Paragraph(report.get("title", "Report"), styles["Title"])]
    meta = report.get("meta", {})
    subtitle = " · ".join(
        f"{key}: {value}" for key, value in meta.items() if not isinstance(value, dict)
    )
    if org_name:
        subtitle = f"{org_name} · {subtitle}"
    story.append(Paragraph(subtitle, styles["Normal"]))
    story.append(Spacer(1, 6))

    header = [Paragraph(f"<b>{c['label']}</b>", small) for c in report["columns"]]
    data = [header]
    for row in report["rows"]:
        data.append(
            [
                Paragraph(
                    str(render_value(row.get(c["key"]), c, duration_format)), small
                )
                for c in report["columns"]
            ]
        )
    if has_column_totals(report):
        data.append(
            [Paragraph(f"<b>{v}</b>", small) for v in totals_row(report, duration_format)]
        )

    if len(data) == 1:
        story.append(Paragraph("No rows matched the selected filters.", styles["Normal"]))
    else:
        table = Table(data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3A5F")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B7C2D0")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -2),
                     [colors.white, colors.HexColor("#F4F6F9")]),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E4E9F0")),
                ]
            )
        )
        story.append(table)

    story.append(Spacer(1, 8))
    story.append(
        Paragraph(
            "Generated by TimeKeeper. Personal data in this report is processed "
            "under the lawful bases documented in the organisation's records of "
            "processing (DP-01).",
            small,
        )
    )
    document.build(story)
    return buffer.getvalue()


MEDIA_TYPES = {
    "csv": "text/csv",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
}


def export(report: dict, fmt: str, duration_format: str = "hm", org_name: str = "") -> bytes:
    if fmt == "csv":
        return to_csv(report, duration_format)
    if fmt == "xlsx":
        return to_xlsx(report, duration_format)
    if fmt == "pdf":
        return to_pdf(report, duration_format, org_name)
    raise ValueError(f"Unsupported export format '{fmt}'")
